# views.py
import io
import json
import logging
from typing import List, Tuple
from .models import (
    AttendanceData,
    TrainingPerformance,
    TrainingPerformanceCategory,
)
from .serializers import (
    StudentAnalyticsSerializer,
)
import openpyxl
from django.db import connection, IntegrityError, DatabaseError, transaction
from django.db.models import Prefetch, Q, Count, Avg, Min, Max, StdDev
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from rest_framework import status, viewsets, views
from rest_framework.decorators import (
    api_view,
    authentication_classes,
    permission_classes,
)
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.authentication import SessionAuthentication, BasicAuthentication
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from base.models import FacultyResponsibility
from student.models import (
    Student,
)

try:
    from .models import TrainingPerformance, TrainingPerformanceCategory
except Exception:
    TrainingPerformance = None
    TrainingPerformanceCategory = None

logger = logging.getLogger(__name__)

TRAINING_CONFIG = {
    "Aptitude": [
        "Arithmetic",
        "Logical Reasoning",
        "Probability",
        "Verbal Ability",
        "Verbal Reasoning",
    ],
    "Technical": ["OS", "DBMS", "DSA", "CN", "OOPS"],
    "Coding": ["Coding Marks"],
}

BASE_HEADERS = ["UID", "Full Name", "Branch"]


def _normalize_header(h):
    return str(h).strip() if h else ""


@csrf_exempt
@api_view(["GET"])
def download_training_template(request, training_type: str):
    training_type = str(training_type).strip()
    if training_type not in TRAINING_CONFIG:
        return JsonResponse(
            {
                "error": "Unknown training_type",
                "valid_types": list(TRAINING_CONFIG.keys()),
            },
            status=400,
        )

    headers = BASE_HEADERS + TRAINING_CONFIG[training_type]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{training_type}_Template"
    ws.append(headers)

    # sample data row
    sample_row = ["101", "John Doe", "CSE_A"] + [
        75 for _ in TRAINING_CONFIG[training_type]
    ]
    ws.append(sample_row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"training_template_{training_type}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


class UploadTrainingPerformanceView(APIView):
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, *args, **kwargs):
        training_type = request.POST.get("training_type")
        semester = request.POST.get("semester")
        date = request.POST.get("date")

        if not training_type:
            return Response(
                {"error": "training_type is required (Aptitude/Technical/Coding)."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not semester:
            return Response(
                {"error": "semester is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        training_type = str(training_type).strip()
        if training_type not in TRAINING_CONFIG:
            return Response(
                {
                    "error": "Unknown training_type",
                    "valid_types": list(TRAINING_CONFIG.keys()),
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        file = request.FILES.get("file")
        if not file:
            return Response(
                {"error": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST
            )

        expected_headers = BASE_HEADERS + TRAINING_CONFIG[training_type]

        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            sheet = wb.active

            header_cells = next(
                sheet.iter_rows(min_row=1, max_row=1, values_only=False)
            )
            read_headers = [_normalize_header(c.value) for c in header_cells]
            read_headers_trim = read_headers[: len(expected_headers)]

            if read_headers_trim != expected_headers:
                return Response(
                    {
                        "error": "Invalid header format.",
                        "expected": expected_headers,
                        "found": read_headers_trim,
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            errors = []
            processed = created_student = updated_student = 0
            created_tp = updated_tp = created_cat = updated_cat = 0
            parsed_rows: List[dict] = []

            for row_idx, row in enumerate(
                sheet.iter_rows(min_row=2, values_only=True), start=2
            ):
                row_list = list(row)
                if len(row_list) < len(expected_headers):
                    row_list += [None] * (len(expected_headers) - len(row_list))

                uid, full_name, branch_div = row_list[:3]
                sub_vals = row_list[3 : 3 + len(TRAINING_CONFIG[training_type])]

                row_errs = []
                if not uid:
                    row_errs.append("UID is required.")
                if not full_name:
                    row_errs.append("Full Name is required.")
                if not branch_div:
                    row_errs.append("Branch is required.")

                sub_marks: List[Tuple[str, float]] = []
                for j, val in enumerate(sub_vals):
                    cat_name = TRAINING_CONFIG[training_type][j]
                    if val is None or str(val).strip() == "":
                        row_errs.append(f"Marks required for '{cat_name}'.")
                    else:
                        try:
                            m = float(val)
                            if m < 0 or m > 100:
                                row_errs.append(
                                    f"Marks out of range (0â€“100) for '{cat_name}'."
                                )
                            sub_marks.append((cat_name, m))
                        except Exception:
                            row_errs.append(f"Marks must be numeric for '{cat_name}'.")

                if row_errs:
                    errors.append({"row": row_idx, "errors": row_errs})
                else:
                    parsed_rows.append(
                        {
                            "row": row_idx,
                            "uid": str(uid).strip(),
                            "full_name": str(full_name).strip(),
                            "branch_div": str(branch_div).strip(),
                            "sub_marks": sub_marks,
                        }
                    )

            uploader_id = request.user.id if request.user.is_authenticated else None

            with transaction.atomic():
                for item in parsed_rows:
                    student_obj, student_created = Student.objects.update_or_create(
                        uid=item["uid"],
                        defaults={
                            "full_name": item["full_name"],
                            "branch_div": item["branch_div"],
                        },
                    )
                    created_student += int(student_created)
                    updated_student += int(not student_created)

                    tp_defaults = {"uploaded_by_id": uploader_id, "date": date}
                    tp_obj, created_flag = TrainingPerformance.objects.update_or_create(
                        student=student_obj,
                        training_type=training_type,
                        semester=semester,
                        defaults=tp_defaults,
                    )
                    created_tp += int(created_flag)
                    updated_tp += int(not created_flag)
                    processed += 1

                    for cat_name, marks in item["sub_marks"]:
                        cat_obj, cat_created = (
                            TrainingPerformanceCategory.objects.update_or_create(
                                performance=tp_obj,
                                category_name=cat_name,
                                defaults={"marks": marks},
                            )
                        )
                        created_cat += int(cat_created)
                        updated_cat += int(not cat_created)

            resp = {
                "message": "Upload processed successfully.",
                "training_type": training_type,
                "semester": semester,
                "processed_rows": processed,
                "created_students": created_student,
                "updated_students": updated_student,
                "created_training_performance": created_tp,
                "updated_training_performance": updated_tp,
                "created_category_rows": created_cat,
                "updated_category_rows": updated_cat,
                "errors_count": len(errors),
                "errors": errors,
            }

            return Response(
                resp,
                status=status.HTTP_201_CREATED
                if len(errors) == 0
                else status.HTTP_207_MULTI_STATUS,
            )

        except openpyxl.utils.exceptions.InvalidFileException:
            return Response(
                {"error": "Invalid Excel file; cannot read."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        except Exception as e:
            logger.exception("Error while uploading training performance file")
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


@api_view(["GET"])
@authentication_classes([SessionAuthentication, BasicAuthentication])
@permission_classes([IsAuthenticated])
def get_attendance_data(request, table_name):
    try:
        valid_tables = [
            "attendance_data",
            "another_table",
        ]  # Add valid table names here
        if table_name not in valid_tables:
            return JsonResponse({"error": "Invalid table name"}, status=400)

        # Construct query with dynamic table name
        query = f"""
            SELECT batch, late, name, present, program_name, session, timestamp, uid, year
            FROM {table_name}
        """

        with connection.cursor() as cursor:
            cursor.execute(query)
            rows = cursor.fetchall()
            # Convert the rows into a list of dictionaries
            columns = [col[0] for col in cursor.description]
            result = [dict(zip(columns, row)) for row in rows]

        return JsonResponse(result, safe=False)

    except Exception as e:
        return JsonResponse(
            {"error": f"Failed to fetch attendance data: {str(e)}"}, status=500
        )


@api_view(["POST"])
@authentication_classes([SessionAuthentication, BasicAuthentication])
@permission_classes([IsAuthenticated])
def save_branch_attendance(request, table_name):
    if request.method == "POST":
        try:
            # Validate table_name to prevent SQL injection
            valid_tables = [
                "batch_attendance",
                "another_batch_table",
            ]  # Add valid table names here
            if table_name not in valid_tables:
                return JsonResponse({"error": "Invalid table name"}, status=400)

            branch_data = request.data.get("branchData")
            if not branch_data:
                return JsonResponse({"error": "No data provided"}, status=400)

            for batch in branch_data:
                batch_name = batch["batch"]
                program_name = batch["program_name"]
                year = batch["year"]
                total_students = batch.get("totalStudents", 0)
                total_present = batch.get("totalPresent", 0)
                total_absent = batch.get("totalAbsent", 0)
                total_late = batch.get("totalLate", 0)

                # Validate each session in batch
                for session, session_data in batch.items():
                    # Skip non-session fields
                    if session in [
                        "batch",
                        "program_name",
                        "year",
                        "totalStudents",
                        "totalPresent",
                        "totalAbsent",
                        "totalLate",
                    ]:
                        continue

                    # Ensure session data has a 'date'
                    if "date" not in session_data:
                        logger.warning(
                            f"Date is missing for session {session} in batch {batch_name}"
                        )
                        return JsonResponse(
                            {"error": f"Date is missing for session {session}"},
                            status=400,
                        )

                    # Get the date from session data
                    session_date = session_data.get("date")

                    # Format the session as "date - session"
                    session_name = f"{session_date} - {session}"

                    # Insert or update batch attendance via raw SQL
                    query = f"""
                        INSERT INTO {table_name} (batch, session, program_name, year, total_students, total_present, total_absent, total_late)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                        ON DUPLICATE KEY UPDATE
                            total_students = VALUES(total_students),
                            total_present = VALUES(total_present),
                            total_absent = VALUES(total_absent),
                            total_late = VALUES(total_late)
                    """
                    try:
                        with connection.cursor() as cursor:
                            cursor.execute(
                                query,
                                (
                                    batch_name,
                                    session_name,  # Use the formatted session name
                                    program_name,
                                    year,
                                    total_students,
                                    total_present,
                                    total_absent,
                                    total_late,
                                ),
                            )
                    except IntegrityError as e:
                        logger.error(
                            f"Integrity error while saving data for batch {batch_name}, session {session_name}: {str(e)}"
                        )
                        return JsonResponse(
                            {"error": f"Database Integrity Error: {str(e)}"}, status=500
                        )
                    except DatabaseError as e:
                        logger.error(
                            f"Database error while saving data for batch {batch_name}, session {session_name}: {str(e)}"
                        )
                        return JsonResponse(
                            {"error": f"Database Error: {str(e)}"}, status=500
                        )
                    except Exception as e:
                        logger.error(
                            f"Error executing query for batch {batch_name}, session {session_name}: {str(e)}"
                        )
                        return JsonResponse(
                            {"error": f"Error executing query: {str(e)}"}, status=500
                        )

            return JsonResponse(
                {
                    "success": True,
                    "message": "Batch-wise attendance saved successfully!",
                }
            )

        except Exception as e:
            logger.exception(
                f"Unexpected error occurred while saving batch attendance: {str(e)}"
            )
            return JsonResponse(
                {"error": f"Failed to save batch-wise attendance: {str(e)}"}, status=500
            )


@api_view(["GET"])
@authentication_classes([SessionAuthentication, BasicAuthentication])
@permission_classes([IsAuthenticated])
def get_avg_data(request, table_name):
    try:
        # Validate table_name to prevent SQL injection
        valid_tables = [
            "program1",
            "another_program_table",
        ]  # Add valid table names here
        if table_name not in valid_tables:
            return JsonResponse({"error": "Invalid table name"}, status=400)

        # Construct query to fetch average attendance and performance by Branch_Div with dynamic table name
        query = f"""
            SELECT Branch_Div,
           Year,
           AVG(training_attendance) AS avg_attendance,
           AVG(training_performance) AS avg_performance
    FROM {table_name}
    GROUP BY Branch_Div, Year
        """

        with connection.cursor() as cursor:
            cursor.execute(query)
            rows = cursor.fetchall()
            # Convert the rows into a list of dictionaries
            columns = [col[0] for col in cursor.description]
            result = [dict(zip(columns, row)) for row in rows]

        return JsonResponse(result, safe=False)

    except Exception as e:
        logger.exception(str(e))
        return JsonResponse(
            {"error": f"Failed to fetch average data: {str(e)}"}, status=500
        )


@api_view(["POST"])
@authentication_classes([SessionAuthentication, BasicAuthentication])
@permission_classes([IsAuthenticated])
def update_attendance(request, table_name):
    if request.method == "POST":
        try:
            # Parse JSON data from request
            data = request.data
            uid = data.get("uid")
            session = data.get("session")
            new_status = data.get("new_status")

            if not uid or not session or new_status not in ["Present", "Absent"]:
                return JsonResponse({"error": "Invalid data provided"}, status=400)

            # Raw SQL query to update attendance
            query = f"""
                UPDATE {table_name}
                SET present = %s
                WHERE uid = %s AND session = %s
            """
            with connection.cursor() as cursor:
                cursor.execute(query, [new_status, uid, session])

            return JsonResponse(
                {"message": "Attendance updated successfully"}, status=200
            )
        except Exception as e:
            logger.exception(str(e))
            return JsonResponse(
                {"error": f"Failed to update attendance: {str(e)}"}, status=500
            )
    else:
        return JsonResponse({"error": "Invalid HTTP method"}, status=405)


class CreateAttendanceRecord(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        data = request.data
        faculty = FacultyResponsibility.objects.get(user=request.user)
        if not faculty.program:
            return Response(
                {"error": "Faculty is not assigned to any program"},
                status=status.HTTP_403_FORBIDDEN,
            )

        # Validate required fields
        required_fields = [
            "phase",
            "semester",
            "year",
            "num_sessions",
            "num_days",
            "dates",
            "file_headers",
            "attendance_data",
        ]
        if not all(field in data for field in required_fields):
            return Response(
                {"error": "Missing required fields"}, status=status.HTTP_400_BAD_REQUEST
            )

        # Check if 'dates', 'file_headers', and 'attendance_data' are lists
        if (
            not isinstance(data["dates"], list)
            or not isinstance(data["file_headers"], list)
            or not isinstance(data["attendance_data"], list)
        ):
            return Response(
                {"error": "Dates, file_headers, and attendance_data must be lists"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Clean the 'attendance_data' to remove unnecessary fields
        cleaned_attendance_data = []
        for record in data["attendance_data"]:
            if "student_data" in record:
                student_data = record["student_data"]
                if isinstance(student_data, list) and len(student_data) == 3:
                    cleaned_attendance_data.append({"student_data": student_data})
                else:
                    return Response(
                        {
                            "error": 'Each "student_data" entry must be a list with 3 elements (UID, Name, Batch)'
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )

        # Prepare the SQL insert statement for attendance record
        program_name = data["program_name"]
        year = data["year"]
        num_sessions = data["num_sessions"]
        semester = data["semester"]
        phase = data["phase"]
        num_days = data["num_days"]
        dates = json.dumps(data["dates"])  # Convert list to JSON string
        file_headers = json.dumps(data["file_headers"])  # Convert list to JSON string
        student_data = json.dumps(
            cleaned_attendance_data
        )  # Convert list of student data to JSON string

        # Create raw SQL insert query
        query = """
            INSERT INTO attendance_attendancerecord (
                program_name, year, num_sessions, num_days, dates, file_headers, student_data, semester, phase
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        # Execute the query using Django's connection cursor
        with connection.cursor() as cursor:
            cursor.execute(
                query,
                [
                    program_name,
                    year,
                    num_sessions,
                    num_days,
                    dates,
                    file_headers,
                    student_data,
                    semester,
                    phase,
                ],
            )

        # Get the ID of the newly inserted record
        record_id = cursor.lastrowid

        # Return success response with created record ID
        return Response(
            {
                "message": "Attendance record successfully created!",
                "record_id": record_id,
            },
            status=status.HTTP_201_CREATED,
        )


class StudentAnalyticsViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = StudentAnalyticsSerializer

    def get_queryset(self):
        queryset = Student.objects.select_related("user").order_by("user__full_name")
        batch = self.request.query_params.get("batch")
        department = self.request.query_params.get("department")
        if batch:
            queryset = queryset.filter(batch=batch)
        if department:
            queryset = queryset.filter(department=department)
        queryset = queryset.prefetch_related(
            Prefetch(
                "trainingperformance_set",
                queryset=TrainingPerformance.objects.prefetch_related("categories"),
            )
        )
        return queryset

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        student_uids = list(queryset.values_list("uid", flat=True))
        attendance_query = AttendanceData.objects.filter(uid__in=student_uids)

        semester = request.query_params.get("semester")
        if semester:
            attendance_query = attendance_query.filter(semester=semester)

        attendance_stats = attendance_query.values("uid").annotate(
            total_sessions=Count("id"),
            present_count=Count("id", filter=Q(present="Present")),
            late_count=Count("id", filter=Q(late="Late")),
        )

        attendance_map = {item["uid"]: item for item in attendance_stats}

        context = {
            "request": request,
            "attendance_map": attendance_map,
        }

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True, context=context)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True, context=context)
        return Response(serializer.data)

class AggregateAnalyticsView(views.APIView):

    def get(self, request, *args, **kwargs):
        batch_filter = request.query_params.get("batch")
        dept_filter = request.query_params.get("department")
        sem_filter = request.query_params.get("semester")
        performance_qs = TrainingPerformance.objects.all()
        attendance_qs = AttendanceData.objects.all()

        if batch_filter:
            performance_qs = performance_qs.filter(student__batch=batch_filter)
            attendance_qs = attendance_qs.filter(batch=batch_filter)
        if dept_filter:
            performance_qs = performance_qs.filter(student__department=dept_filter)
            attendance_qs = attendance_qs.filter(program_name=dept_filter)
        if sem_filter:
            performance_qs = performance_qs.filter(semester=sem_filter)
            attendance_qs = attendance_qs.filter(semester=sem_filter)
        perf_group_fields = ["student__batch", "student__department", "semester"]
        att_group_fields = ["batch", "program_name", "semester"]

        performance_overview_aggs = (
            performance_qs.values(*perf_group_fields)
            .annotate(
                average_score=Avg("categories__marks"),
                min_score=Min("categories__marks"),
                max_score=Max("categories__marks"),
                std_dev_score=StdDev("categories__marks"),
                total_students_tested=Count("student", distinct=True)
            )
            .order_by(*perf_group_fields)
        )

        perf_overview_map = {
            (item["student__batch"], item["student__department"], item["semester"]): {
                "average_score": round(item["average_score"], 2) if item["average_score"] else 0,
                "min_score": item["min_score"] or 0,
                "max_score": item["max_score"] or 0,
                "std_dev_score": round(item["std_dev_score"], 2) if item["std_dev_score"] else 0,
                "total_students_tested": item["total_students_tested"],
            }
            for item in performance_overview_aggs
        }

        category_aggs = (
            performance_qs.values(*perf_group_fields, "categories__category_name")
            .annotate(average_score=Avg("categories__marks"))
            .filter(categories__category_name__isnull=False)
            .order_by(*perf_group_fields, "categories__category_name")
        )
        category_map = {}
        for item in category_aggs:
            group_key = (item["student__batch"], item["student__department"], item["semester"])
            category_name = item["categories__category_name"]
            avg_score = round(item["average_score"], 2) if item["average_score"] else 0

            if group_key not in category_map:
                category_map[group_key] = {}

            category_map[group_key][category_name] = avg_score
        attendance_aggs = (
            attendance_qs.values(*att_group_fields)
            .annotate(
                total_sessions=Count("id"),
                present_count=Count("id", filter=Q(present="Present")),
                late_count=Count("id", filter=Q(late="Late")),
                total_students_attended=Count("uid", distinct=True)
            )
            .order_by(*att_group_fields)
        )

        attendance_map = {
            (item["batch"], item["program_name"], item["semester"]): item
            for item in attendance_aggs
        }
        all_keys = (
            set(perf_overview_map.keys()) |
            set(category_map.keys()) |
            set(attendance_map.keys())
        )

        final_data = []
        for group_key in all_keys:
            batch, department, semester = group_key

            if batch is None or department is None or semester is None:
                continue

            att_data = attendance_map.get(group_key)
            perf_overview_stats = perf_overview_map.get(group_key, {})
            category_stats = category_map.get(group_key, {})
            if att_data:
                total = att_data.get("total_sessions", 0)
                present = att_data.get("present_count", 0)
                late = att_data.get("late_count", 0)
                absent = total - present

                attendance_summary = {
                    "average_attendance": round((present / total) * 100, 2) if total > 0 else 0,
                    "late_percentage": round((late / total) * 100, 2) if total > 0 else 0,
                    "absent_percentage": round((absent / total) * 100, 2) if total > 0 else 0,
                    "total_students_attended": att_data.get("total_students_attended", 0),
                    "total_sessions": total,
                    "late_sessions": late,
                    "absent_sessions": absent,
                }
            else:
                attendance_summary = {}

            final_data.append({
                "batch": batch,
                "department": department,
                "semester": semester,
                "attendance_summary": attendance_summary,
                "performance_overview": perf_overview_stats,
                "performance_by_category": category_stats
            })

        return Response(final_data)